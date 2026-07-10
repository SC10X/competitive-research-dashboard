#!/usr/bin/env python3
"""导出全部品牌数据为多Sheet Excel文件"""

import json
import sqlite3
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = "backend/data/competitive_research.db"
OUTPUT = "/workspace/北美服饰鞋包竞对品牌研究看板_完整数据.xlsx"

def safe_json(v):
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)

def safe_date(v):
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)[:10]

def fmt_num(v, decimals=1):
    if v is None:
        return ""
    try:
        return round(float(v), decimals)
    except (ValueError, TypeError):
        return v

def fmt_bool(v):
    if v is None:
        return ""
    return "是" if v else "否"

def style_header(ws, headers, col_widths=None):
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

def style_data(ws, row_start, row_end, col_count):
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for r in range(row_start, row_end + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if (r - row_start) % 2 == 1:
                cell.fill = alt_fill

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
c = conn.cursor()

wb = Workbook()

# ========== Sheet 1: 品牌概览 ==========
ws1 = wb.active
ws1.title = "品牌概览"

# 预加载分类映射
c.execute("SELECT brand_id, category_id, is_primary FROM brand_categories")
bc_rows = c.fetchall()
cat_map = {}
c.execute("SELECT id, name, slug FROM categories")
cat_rows = c.fetchall()
cat_id_name = {r["id"]: r["name"] for r in cat_rows}
cat_id_slug = {r["id"]: r["slug"] for r in cat_rows}
for bc in bc_rows:
    bid = bc["brand_id"]
    if bid not in cat_map:
        cat_map[bid] = {"primary": [], "all": []}
    name = cat_id_name.get(bc["category_id"], "")
    if bc["is_primary"]:
        cat_map[bid]["primary"].append(name)
    cat_map[bid]["all"].append(name)

c.execute("SELECT * FROM brands ORDER BY id")
brands = c.fetchall()

headers1 = ["ID", "品牌名称", "Slug", "国家", "成立年份", "母公司", "总部", "官网", "描述",
            "状态", "主要分类", "所有分类"]
col_widths1 = [6, 30, 25, 12, 10, 20, 25, 30, 60, 8, 20, 30]
style_header(ws1, headers1, col_widths1)

for i, b in enumerate(brands):
    row = i + 2
    cats = cat_map.get(b["id"], {"primary": [], "all": []})
    ws1.cell(row=row, column=1, value=b["id"])
    ws1.cell(row=row, column=2, value=b["name"])
    ws1.cell(row=row, column=3, value=b["slug"])
    ws1.cell(row=row, column=4, value=b["country"])
    ws1.cell(row=row, column=5, value=b["founded_year"])
    ws1.cell(row=row, column=6, value=b["parent_company"])
    ws1.cell(row=row, column=7, value=b["headquarters"])
    ws1.cell(row=row, column=8, value=b["website"])
    ws1.cell(row=row, column=9, value=b["description"])
    ws1.cell(row=row, column=10, value="活跃" if b["is_active"] else "停用")
    ws1.cell(row=row, column=11, value=", ".join(cats["primary"]))
    ws1.cell(row=row, column=12, value=", ".join(cats["all"]))

style_data(ws1, 2, len(brands) + 1, len(headers1))

# ========== Sheet 2: 品牌定位 ==========
ws2 = wb.create_sheet("品牌定位")
c.execute("""SELECT bp.*, b.name as brand_name FROM brand_positioning bp 
             JOIN brands b ON bp.brand_id = b.id ORDER BY b.id""")
rows2 = c.fetchall()
headers2 = ["ID", "品牌名称", "价格带", "品牌调性", "核心价值主张", "USP", "品牌故事",
            "数据来源URL", "来源类型", "最后验证日期"]
col_widths2 = [6, 30, 12, 15, 50, 50, 60, 40, 12, 14]
style_header(ws2, headers2, col_widths2)
for i, r in enumerate(rows2):
    row = i + 2
    ws2.cell(row=row, column=1, value=r["brand_id"])
    ws2.cell(row=row, column=2, value=r["brand_name"])
    ws2.cell(row=row, column=3, value=r["price_tier"])
    ws2.cell(row=row, column=4, value=r["brand_tone"])
    ws2.cell(row=row, column=5, value=r["core_value_proposition"])
    ws2.cell(row=row, column=6, value=r["usp"])
    ws2.cell(row=row, column=7, value=r["brand_story"])
    ws2.cell(row=row, column=8, value=r["source_url"])
    ws2.cell(row=row, column=9, value=r["source_type"])
    ws2.cell(row=row, column=10, value=safe_date(r["last_verified_at"]))
style_data(ws2, 2, len(rows2) + 1, len(headers2))

# ========== Sheet 3: 定价策略 ==========
ws3 = wb.create_sheet("定价策略")
c.execute("""SELECT ps.*, b.name as brand_name FROM pricing_strategy ps
             JOIN brands b ON ps.brand_id = b.id ORDER BY b.id, ps.id""")
rows3 = c.fetchall()
headers3 = ["ID", "品牌名称", "品类", "最低价($)", "最高价($)", "均价($)",
            "折扣频率", "典型折扣%", "有会员", "会员价($)", "会员模式", "数据日期",
            "来源URL", "来源类型"]
col_widths3 = [6, 30, 15, 12, 12, 10, 12, 12, 8, 12, 25, 12, 40, 12]
style_header(ws3, headers3, col_widths3)
for i, r in enumerate(rows3):
    row = i + 2
    ws3.cell(row=row, column=1, value=r["brand_id"])
    ws3.cell(row=row, column=2, value=r["brand_name"])
    ws3.cell(row=row, column=3, value=r["category_name"])
    ws3.cell(row=row, column=4, value=fmt_num(r["price_range_min"]))
    ws3.cell(row=row, column=5, value=fmt_num(r["price_range_max"]))
    ws3.cell(row=row, column=6, value=fmt_num(r["avg_price"]))
    ws3.cell(row=row, column=7, value=r["discount_frequency"])
    ws3.cell(row=row, column=8, value=fmt_num(r["typical_discount_pct"]))
    ws3.cell(row=row, column=9, value=fmt_bool(r["has_membership"]))
    ws3.cell(row=row, column=10, value=fmt_num(r["membership_price"]))
    ws3.cell(row=row, column=11, value=r["membership_model"])
    ws3.cell(row=row, column=12, value=safe_date(r["data_as_of"]))
    ws3.cell(row=row, column=13, value=r["source_url"])
    ws3.cell(row=row, column=14, value=r["source_type"])
style_data(ws3, 2, len(rows3) + 1, len(headers3))

# ========== Sheet 4: 目标人群 ==========
ws4 = wb.create_sheet("目标人群")
c.execute("""SELECT td.*, b.name as brand_name FROM target_demographics td
             JOIN brands b ON td.brand_id = b.id ORDER BY b.id""")
rows4 = c.fetchall()
headers4 = ["ID", "品牌名称", "年龄段", "性别倾向", "收入水平", "生活方式标签",
            "核心场景", "来源URL", "来源类型"]
col_widths4 = [6, 30, 12, 12, 12, 40, 40, 40, 12]
style_header(ws4, headers4, col_widths4)
for i, r in enumerate(rows4):
    row = i + 2
    ws4.cell(row=row, column=1, value=r["brand_id"])
    ws4.cell(row=row, column=2, value=r["brand_name"])
    ws4.cell(row=row, column=3, value=r["age_range"])
    ws4.cell(row=row, column=4, value=r["gender_skew"])
    ws4.cell(row=row, column=5, value=r["income_level"])
    ws4.cell(row=row, column=6, value=safe_json(r["lifestyle_tags"]))
    ws4.cell(row=row, column=7, value=safe_json(r["core_scenarios"]))
    ws4.cell(row=row, column=8, value=r["source_url"])
    ws4.cell(row=row, column=9, value=r["source_type"])
style_data(ws4, 2, len(rows4) + 1, len(headers4))

# ========== Sheet 5: 产品策略 ==========
ws5 = wb.create_sheet("产品策略")
c.execute("""SELECT pds.*, b.name as brand_name FROM product_strategy pds
             JOIN brands b ON pds.brand_id = b.id ORDER BY b.id""")
rows5 = c.fetchall()
headers5 = ["ID", "品牌名称", "英雄产品", "发布节奏", "SKU数(估)", "可持续产品线",
            "联名策略", "近期联名", "技术创新", "品类扩展", "数据日期", "来源URL", "来源类型"]
col_widths5 = [6, 30, 40, 15, 12, 12, 30, 40, 40, 40, 12, 40, 12]
style_header(ws5, headers5, col_widths5)
for i, r in enumerate(rows5):
    row = i + 2
    ws5.cell(row=row, column=1, value=r["brand_id"])
    ws5.cell(row=row, column=2, value=r["brand_name"])
    ws5.cell(row=row, column=3, value=safe_json(r["hero_products"]))
    ws5.cell(row=row, column=4, value=r["launch_cadence"])
    ws5.cell(row=row, column=5, value=r["sku_count_estimate"])
    ws5.cell(row=row, column=6, value=fmt_bool(r["has_sustainable_line"]))
    ws5.cell(row=row, column=7, value=r["collab_strategy"])
    ws5.cell(row=row, column=8, value=safe_json(r["recent_collabs"]))
    ws5.cell(row=row, column=9, value=safe_json(r["tech_innovations"]))
    ws5.cell(row=row, column=10, value=safe_json(r["category_expansion"]))
    ws5.cell(row=row, column=11, value=safe_date(r["data_as_of"]))
    ws5.cell(row=row, column=12, value=r["source_url"])
    ws5.cell(row=row, column=13, value=r["source_type"])
style_data(ws5, 2, len(rows5) + 1, len(headers5))

# ========== Sheet 6: 渠道策略 ==========
ws6 = wb.create_sheet("渠道策略")
c.execute("""SELECT cs.*, b.name as brand_name FROM channel_strategy cs
             JOIN brands b ON cs.brand_id = b.id ORDER BY b.id""")
rows6 = c.fetchall()
headers6 = ["ID", "品牌名称", "DTC占比%", "批发占比%", "北美自营店数",
            "零售合作伙伴", "电商平台", "国际市场", "数据日期", "来源URL", "来源类型"]
col_widths6 = [6, 30, 12, 12, 14, 40, 30, 30, 12, 40, 12]
style_header(ws6, headers6, col_widths6)
for i, r in enumerate(rows6):
    row = i + 2
    ws6.cell(row=row, column=1, value=r["brand_id"])
    ws6.cell(row=row, column=2, value=r["brand_name"])
    ws6.cell(row=row, column=3, value=fmt_num(r["dtc_pct"]))
    ws6.cell(row=row, column=4, value=fmt_num(r["wholesale_pct"]))
    ws6.cell(row=row, column=5, value=r["own_stores_na"])
    ws6.cell(row=row, column=6, value=safe_json(r["retail_partners"]))
    ws6.cell(row=row, column=7, value=safe_json(r["ecommerce_platforms"]))
    ws6.cell(row=row, column=8, value=safe_json(r["international_markets"]))
    ws6.cell(row=row, column=9, value=safe_date(r["data_as_of"]))
    ws6.cell(row=row, column=10, value=r["source_url"])
    ws6.cell(row=row, column=11, value=r["source_type"])
style_data(ws6, 2, len(rows6) + 1, len(headers6))

# ========== Sheet 7: 社交媒体 ==========
ws7 = wb.create_sheet("社交媒体")
c.execute("""SELECT sm.*, b.name as brand_name FROM social_media_metrics sm
             JOIN brands b ON sm.brand_id = b.id ORDER BY b.id, sm.id""")
rows7 = c.fetchall()
headers7 = ["ID", "品牌名称", "平台", "粉丝数", "互动率%", "平均点赞", "平均评论",
            "发布频率", "热门标签", "关键KOL", "数据日期", "来源URL", "来源类型"]
col_widths7 = [6, 30, 12, 14, 10, 12, 12, 12, 30, 30, 12, 40, 12]
style_header(ws7, headers7, col_widths7)
for i, r in enumerate(rows7):
    row = i + 2
    ws7.cell(row=row, column=1, value=r["brand_id"])
    ws7.cell(row=row, column=2, value=r["brand_name"])
    ws7.cell(row=row, column=3, value=r["platform"])
    ws7.cell(row=row, column=4, value=r["followers"])
    ws7.cell(row=row, column=5, value=fmt_num(r["engagement_rate"]))
    ws7.cell(row=row, column=6, value=r["avg_likes"])
    ws7.cell(row=row, column=7, value=r["avg_comments"])
    ws7.cell(row=row, column=8, value=r["post_frequency"])
    ws7.cell(row=row, column=9, value=safe_json(r["top_hashtags"]))
    ws7.cell(row=row, column=10, value=safe_json(r["key_kols"]))
    ws7.cell(row=row, column=11, value=safe_date(r["data_as_of"]))
    ws7.cell(row=row, column=12, value=r["source_url"])
    ws7.cell(row=row, column=13, value=r["source_type"])
style_data(ws7, 2, len(rows7) + 1, len(headers7))

# ========== Sheet 8: 数字化能力 ==========
ws8 = wb.create_sheet("数字化能力")
c.execute("""SELECT dc.*, b.name as brand_name FROM digital_capability dc
             JOIN brands b ON dc.brand_id = b.id ORDER BY b.id""")
rows8 = c.fetchall()
headers8 = ["ID", "品牌名称", "月网站访问量", "App评分", "App下载量", "个性化推荐",
            "虚拟试穿", "社区功能", "会员计划", "会员名称", "AI功能", "数据日期",
            "来源URL", "来源类型"]
col_widths8 = [6, 30, 16, 10, 14, 12, 10, 10, 10, 20, 30, 12, 40, 12]
style_header(ws8, headers8, col_widths8)
for i, r in enumerate(rows8):
    row = i + 2
    ws8.cell(row=row, column=1, value=r["brand_id"])
    ws8.cell(row=row, column=2, value=r["brand_name"])
    ws8.cell(row=row, column=3, value=r["monthly_web_visits"])
    ws8.cell(row=row, column=4, value=r["app_rating"])
    ws8.cell(row=row, column=5, value=r["app_downloads"])
    ws8.cell(row=row, column=6, value=fmt_bool(r["has_personalization"]))
    ws8.cell(row=row, column=7, value=fmt_bool(r["has_virtual_tryon"]))
    ws8.cell(row=row, column=8, value=fmt_bool(r["has_community_feature"]))
    ws8.cell(row=row, column=9, value=fmt_bool(r["has_membership_program"]))
    ws8.cell(row=row, column=10, value=r["membership_name"])
    ws8.cell(row=row, column=11, value=safe_json(r["ai_features"]))
    ws8.cell(row=row, column=12, value=safe_date(r["data_as_of"]))
    ws8.cell(row=row, column=13, value=r["source_url"])
    ws8.cell(row=row, column=14, value=r["source_type"])
style_data(ws8, 2, len(rows8) + 1, len(headers8))

# ========== Sheet 9: 财务表现 ==========
ws9 = wb.create_sheet("财务表现")
c.execute("""SELECT fp.*, b.name as brand_name FROM financial_performance fp
             JOIN brands b ON fp.brand_id = b.id ORDER BY b.id, fp.fiscal_year DESC""")
rows9 = c.fetchall()
headers9 = ["ID", "品牌名称", "财年", "营收($M)", "营收增速%", "毛利率%",
            "北美营收占比%", "估算", "数据来源"]
col_widths9 = [6, 30, 8, 14, 12, 12, 14, 8, 30]
style_header(ws9, headers9, col_widths9)
for i, r in enumerate(rows9):
    row = i + 2
    ws9.cell(row=row, column=1, value=r["brand_id"])
    ws9.cell(row=row, column=2, value=r["brand_name"])
    ws9.cell(row=row, column=3, value=r["fiscal_year"])
    ws9.cell(row=row, column=4, value=fmt_num(r["revenue"]))
    ws9.cell(row=row, column=5, value=fmt_num(r["revenue_growth_pct"]))
    ws9.cell(row=row, column=6, value=fmt_num(r["gross_margin_pct"]))
    ws9.cell(row=row, column=7, value=fmt_num(r["na_revenue_pct"]))
    ws9.cell(row=row, column=8, value="估算" if r["is_estimated"] else "实际")
    ws9.cell(row=row, column=9, value=r["data_source"])
style_data(ws9, 2, len(rows9) + 1, len(headers9))

# ========== Sheet 10: 用户评价 ==========
ws10 = wb.create_sheet("用户评价")
c.execute("""SELECT cs.*, b.name as brand_name FROM customer_sentiment cs
             JOIN brands b ON cs.brand_id = b.id ORDER BY b.id, cs.id""")
rows10 = c.fetchall()
headers10 = ["ID", "品牌名称", "平台", "评分", "评价数", "正面主题",
             "负面主题", "NPS", "数据日期", "来源URL", "来源类型"]
col_widths10 = [6, 30, 12, 8, 10, 40, 40, 8, 12, 40, 12]
style_header(ws10, headers10, col_widths10)
for i, r in enumerate(rows10):
    row = i + 2
    ws10.cell(row=row, column=1, value=r["brand_id"])
    ws10.cell(row=row, column=2, value=r["brand_name"])
    ws10.cell(row=row, column=3, value=r["platform"])
    ws10.cell(row=row, column=4, value=r["rating"])
    ws10.cell(row=row, column=5, value=r["review_count"])
    ws10.cell(row=row, column=6, value=safe_json(r["positive_themes"]))
    ws10.cell(row=row, column=7, value=safe_json(r["negative_themes"]))
    ws10.cell(row=row, column=8, value=r["nps_score"])
    ws10.cell(row=row, column=9, value=safe_date(r["data_as_of"]))
    ws10.cell(row=row, column=10, value=r["source_url"])
    ws10.cell(row=row, column=11, value=r["source_type"])
style_data(ws10, 2, len(rows10) + 1, len(headers10))

# ========== Sheet 11: 竞对事件 ==========
ws11 = wb.create_sheet("竞对事件")
c.execute("""SELECT ce.*, b.name as brand_name FROM competitive_events ce
             JOIN brands b ON ce.brand_id = b.id ORDER BY ce.event_date DESC, b.id""")
rows11 = c.fetchall()
headers11 = ["ID", "品牌名称", "事件类型", "标题", "描述", "事件日期",
             "重要性", "标签", "来源名称", "来源URL", "原文引用"]
col_widths11 = [6, 30, 16, 50, 80, 12, 10, 25, 20, 50, 60]
style_header(ws11, headers11, col_widths11)
for i, r in enumerate(rows11):
    row = i + 2
    ws11.cell(row=row, column=1, value=r["id"])
    ws11.cell(row=row, column=2, value=r["brand_name"])
    ws11.cell(row=row, column=3, value=r["event_type"])
    ws11.cell(row=row, column=4, value=r["title"])
    ws11.cell(row=row, column=5, value=r["description"])
    ws11.cell(row=row, column=6, value=safe_date(r["event_date"]))
    ws11.cell(row=row, column=7, value=r["importance"])
    ws11.cell(row=row, column=8, value=safe_json(r["tags"]))
    ws11.cell(row=row, column=9, value=r["source_name"])
    ws11.cell(row=row, column=10, value=r["source_url"])
    ws11.cell(row=row, column=11, value=r["source_quote"])
style_data(ws11, 2, len(rows11) + 1, len(headers11))

# ========== Sheet 12: 品牌分类 ==========
ws12 = wb.create_sheet("品牌分类")
c.execute("""SELECT bc.*, b.name as brand_name, c.name as cat_name, c.slug as cat_slug
             FROM brand_categories bc
             JOIN brands b ON bc.brand_id = b.id
             JOIN categories c ON bc.category_id = c.id
             ORDER BY b.id, bc.is_primary DESC""")
rows12 = c.fetchall()
headers12 = ["品牌ID", "品牌名称", "分类ID", "分类名称", "分类Slug", "是否主分类"]
col_widths12 = [8, 30, 10, 25, 25, 12]
style_header(ws12, headers12, col_widths12)
for i, r in enumerate(rows12):
    row = i + 2
    ws12.cell(row=row, column=1, value=r["brand_id"])
    ws12.cell(row=row, column=2, value=r["brand_name"])
    ws12.cell(row=row, column=3, value=r["category_id"])
    ws12.cell(row=row, column=4, value=r["cat_name"])
    ws12.cell(row=row, column=5, value=r["cat_slug"])
    ws12.cell(row=row, column=6, value="是" if r["is_primary"] else "否")
style_data(ws12, 2, len(rows12) + 1, len(headers12))

# 保存
wb.save(OUTPUT)
conn.close()

print(f"✅ 导出完成！文件: {OUTPUT}")
print(f"共 12 个Sheet:")
sheets_info = [
    ("品牌概览", len(brands), "469个品牌基本信息+分类"),
    ("品牌定位", len(rows2), "价格带/调性/价值主张/USP/品牌故事"),
    ("定价策略", len(rows3), "各品类价格区间/折扣/会员信息"),
    ("目标人群", len(rows4), "年龄/性别/收入/生活方式/场景"),
    ("产品策略", len(rows5), "英雄产品/联名/技术创新/品类扩展"),
    ("渠道策略", len(rows6), "DTC/批发占比/门店/电商/国际市场"),
    ("社交媒体", len(rows7), "各平台粉丝/互动率/KOL"),
    ("数字化能力", len(rows8), "网站流量/App/AI功能/会员"),
    ("财务表现", len(rows9), "营收/增速/毛利率"),
    ("用户评价", len(rows10), "评分/评价数/NPS/正负面主题"),
    ("竞对事件", len(rows11), "424条事件含原文引用"),
    ("品牌分类", len(rows12), "品牌-分类关联表"),
]
for name, cnt, desc in sheets_info:
    print(f"  📊 {name}: {cnt} 行 — {desc}")
